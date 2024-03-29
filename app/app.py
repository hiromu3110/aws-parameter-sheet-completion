import copy
import json
import logging
import re
import sys
import boto3
import openpyxl
import jsonpath_ng.ext

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

session = None
clients = dict()
upper_case = re.compile(r'([A-Z])')


# converts a string from camel case to snake case
def to_snake(camel):
    return upper_case.sub(r'_\1', camel)[1:].lower()


# checks whether the action is non-destructive
def is_safe_action(action_name):
    if action_name.startswith('Get'):
        return True
    if action_name.startswith('Describe'):
        return True
    if action_name.startswith('List'):
        return True
    return False


# fetch a single value with JSONPath
def get_value(json, path):
    values = get_values(json, path)
    if 0 < len(values):
        # NOTE: returns the first element
        return values[0]
    else:
        return None


# fetch multiple values with JSONPath
def get_values(json_string, path):
    logger.info('JSONPath string: ' + path)
    values = [x.value for x in jsonpath_ng.ext.parse(path).find(json_string)]
    logger.info('Parsed values: ' + json.dumps(values))
    return values


# resolves the placeholders
def resolve_placeholders(template, symbol, start_cell):
    logger.debug(f'{start_cell.coordinate}: template string is {template}')
    result = template
    cell = start_cell
    for i in range(1, 1 + template.count(symbol)):
        # e.g., %1 -> foo, %2 -> bar
        result = result.replace(f'{symbol}{i}', cell.value)
        cell = cell.offset(column=1)
    logger.debug(f'{cell.coordinate}: resolved string is {result}')
    return result


# calls the API
def invoke(api_name, region_name, action_name, request_params):
    client = clients.get((api_name, region_name))
    if client is None:
        try:
            client = session.client(api_name, region_name=region_name)
            clients[(api_name, region_name)] = client
        except Exception as e:
            logger.error('API name or region name is not valid')
            logger.error('API name: ' + api_name)
            logger.error('region name: ' + region_name)
            raise e

    try:
        method_name = to_snake(action_name)
        method = getattr(client, method_name)
    except Exception as e:
        logger.error('no such action: ' + action_name)
        raise e

    if not is_safe_action(action_name):
        raise Exception('Performing unsafe action was detected: ' +
                        action_name)

    try:
        request = json.loads(request_params)
    except Exception as e:
        logger.error('parameter string is not valid JSON: ' + request_params)
        raise e

    try:
        logger.info('API: ' + api_name)
        logger.info('Region: ' + region_name)
        logger.info('Action: ' + action_name)
        logger.info('Request: ' + json.dumps(request, indent=2))
        response = method(**request)
        logger.info('Response: ' + json.dumps(response, indent=2, default=str))
        return response
    except Exception as e:
        logger.error('request is not accepted: ' +
                     json.dumps(request, indent=2))
        raise e


# reads API parameters
def read_api_params(cell, args):
    ws = cell.parent
    for row in ws.iter_rows(values_only=True,
                            min_row=cell.row, max_row=cell.row,
                            min_col=cell.column, max_col=cell.column+3):
        api_name, region_name, action_name, req_params = row

    logger.debug(f'{cell.row}: API name is [{api_name}]')
    logger.debug(f'{cell.row}: region name is [{region_name}]')
    logger.debug(f'{cell.row}: action name is [{action_name}]')
    logger.debug(f'{cell.row}: request params is {req_params}')

    symbol = '%'
    for i, x in enumerate(args):
        req_params = req_params.replace(f'{symbol}{i + 1}', x)
    logger.debug(f'resolved request params is {req_params}')

    return dict(api_name=api_name,
                region_name=region_name,
                action_name=action_name,
                request_params=req_params)


# reads the row and builds a JSONPath string
def read_path(start_cell):
    # if a row is [foo, bar, baz], this method returns $.foo.bar.baz
    string = ''

    cell = start_cell
    while cell.value:
        string += '.' + cell.value
        cell = cell.offset(column=1)

    if string:
        logger.debug('JSONPath: ' + string)
        return '$' + string
    else:
        raise Exception('Path is not found: ' + start_cell.coordinate)


# finds a symbol in the row
def find_column_symbol(symbol, start_cell):
    max_column = start_cell.parent.max_column

    cell = start_cell
    while cell.column <= max_column:
        logger.debug(f'seeking [{symbol}] in {cell.coordinate}')
        if cell.value == symbol:
            logger.debug(f'[{symbol}] is found in {cell.coordinate}')
            return cell
        cell = cell.offset(column=1)
    logger.info(f'[{symbol}] is not found in {cell.coordinate}')
    raise Exception(f'Symbol [{symbol}] is not found: ' +
                    start_cell.column_letter)


# finds a cell to process
def find_next_cell(symbol, start_cell):
    cell = find_column_symbol(symbol, start_cell)
    return cell.offset(column=1)


# writes a value
def write_value(cell, value):
    if value is None:
        cell.number_format = openpyxl.styles.numbers.FORMAT_GENERAL
        cell.value = '=NA()'
    else:
        cell.number_format = openpyxl.styles.numbers.FORMAT_TEXT
        cell.value = str(value)


def find_form(ws):
    top = None
    right = None

    for row in ws.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        logger.debug(f'current row is {cell.row}')
        if cell.value == '%top':
            top = cell.row
            left_cell = find_column_symbol('%left', cell)
            right_cell = find_column_symbol('%right', left_cell)
            left = left_cell.column
            right = right_cell.column
        elif cell.value == '%bottom':
            bottom = cell.row
            return top, bottom, left, right

    if not top:
        raise Exception('%top is not found.')
    else:
        raise Exception('%bottom is not found.')


def copy_column_dimensions(dimensions, src_col, dst_col):
    from openpyxl.utils import get_column_letter

    dim = dimensions.get(get_column_letter(src_col))
    if dim is None:
        return
    dimensions[get_column_letter(dst_col)].width = dim.width
    dimensions[get_column_letter(dst_col)].hidden = dim.hidden
    if dim.outline_level != 0:
        dimensions.group(get_column_letter(dst_col),
                         get_column_letter(dst_col + dim.max - dim.min),
                         outline_level=dim.outline_level,
                         hidden=dim.hidden)


def copy_form(ws, count):
    from openpyxl.worksheet.cell_range import CellRange

    top, bottom, left, right = find_form(ws)

    src_range = CellRange(min_row=top, max_row=bottom,
                          min_col=left, max_col=right)
    src_merged_cell_ranges = [r for r in ws.merged_cells.ranges
                              if src_range.issuperset(r)]

    ws.sheet_properties.outlinePr.summaryRight = False

    work_col = right + 1
    for i in range(1, count + 1):
        dst_range = CellRange(min_row=top, max_row=bottom, min_col=work_col,
                              max_col=(work_col + right - left))
        for r in ws.merged_cells.ranges:
            if not dst_range.isdisjoint(r):
                ws.unmerge_cells(r.coord)
        for j, col in enumerate(ws.iter_cols(min_row=top, max_row=bottom,
                                             min_col=left, max_col=right)):
            copy_column_dimensions(ws.column_dimensions, left + j, work_col)
            for cell in col:
                dst_cell = ws.cell(row=cell.row, column=work_col)
                if cell.has_style:
                    dst_cell._style = cell._style
                if cell.row == top and cell.value == '%left':
                    dst_cell.value = f'%left{i}'
                elif cell.row == top and cell.value == '%right':
                    dst_cell.value = f'%right{i}'
                elif cell.data_type == 'f':
                    dst_cell.value = openpyxl.formula.translate.Translator(
                        cell.value, origin=cell.coordinate
                    ).translate_formula(dst_cell.coordinate)
                else:
                    dst_cell.value = cell.value
            work_col += 1
        for r in src_merged_cell_ranges:
            copied = copy.copy(r)
            copied.shift(col_shift=(work_col - right - 1))
            ws.merge_cells(copied.coord)

    for row in ws.iter_rows(min_col=work_col, max_row=bottom):
        for cell in row:
            cell.value = None


def process_worksheet(ws, args):
    copy_form(ws, len(args))

    response = None
    for i, x in enumerate(args):
        left = None
        for row in ws.iter_rows(min_col=1, max_col=1):
            cell = row[0]
            logger.debug(f'current row is {cell.row}')
            if cell.value == '%top':
                left = find_column_symbol(f'%left{i + 1}', cell)
            elif cell.value == '%bottom':
                break
            elif left and cell.value == '#call':
                input_cell = find_next_cell('##', cell)
                request = read_api_params(input_cell, x)
                response = invoke(**request)
            elif left and cell.value == '#output':
                input_cell = find_next_cell('##', cell)
                path = read_path(input_cell)
                if '%' in path:
                    param_cell = find_next_cell('###', cell)
                    path = resolve_placeholders(path, '%', param_cell)
                value = get_value(response, path)
                output_cell = find_next_cell('####',
                                             ws.cell(cell.row, left.column))
                write_value(output_cell, value)


def read_target_resources_by_sheet(ws):
    sheet_names = ws.parent.sheetnames
    result = dict()
    for row in ws.iter_rows(values_only=True, min_col=1):
        sheet_name = row[0]
        if not sheet_name:
            continue
        if sheet_name not in sheet_names:
            logger.error('no such worksheet: ' + sheet_name)
            continue
        result.setdefault(sheet_name, [])
        args = [x for x in row[1:] if x is not None]
        result[sheet_name].append(args)
    return result


def process_workbook(src_filename, dst_filename):
    wb = openpyxl.load_workbook(src_filename)
    ws = wb['TargetResources']
    try:
        sheets = read_target_resources_by_sheet(ws)
        for sheet_name, args in sheets.items():
            process_worksheet(wb[sheet_name], args)
        wb.save(dst_filename)
    except Exception as e:
        raise Exception('Failed: ' + sheet_name) from e


if __name__ == '__main__':
    sh = logging.StreamHandler(sys.stdout)
    logger.addHandler(sh)

    if len(sys.argv) < 2:
        raise Exception('Excel book is not specified.')
    src_file = sys.argv[1]
    if not src_file.lower().endswith('.xlsx'):
        raise Exception("Specified file's extension is not XLSX: " + src_file)
    dst_file = re.sub(r'\.xlsx$', '_.xlsx', src_file, flags=re.IGNORECASE)

    if 2 < len(sys.argv):
        profile_name = sys.argv[2]
    else:
        profile_name = 'default'
    session = boto3.Session(profile_name=profile_name)
    identity = session.client('sts').get_caller_identity()
    logger.info('GetCallerIdentity: ' + json.dumps(identity, indent=2))

    process_workbook(src_file, dst_file)
